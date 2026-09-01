import csv
import hashlib
import io
import json
import uuid
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from fastapi import HTTPException, status
from sqlmodel import Session, select

from app.api.attendee.models import Attendees
from app.api.custom_export.registry import (
    DATASETS,
    ExportDatasetDefinition,
    ExportFieldDefinition,
)
from app.api.custom_export.schemas import (
    CustomExportSpec,
    ExportCatalogPublic,
    ExportColumn,
    ExportDatasetPublic,
    ExportFieldPublic,
    ExportFilter,
    ExportFormat,
    ExportPreview,
    ExportPreviewColumn,
)
from app.utils.tabular_export import safe_tabular_cell

MAX_SOURCE_ROWS = 50_000
MAX_EXPORT_ROWS = 50_000
MAX_COLUMNS = 25


def export_catalog() -> ExportCatalogPublic:
    return ExportCatalogPublic(
        datasets=[
            ExportDatasetPublic(
                dataset=dataset.dataset,
                label=dataset.label,
                description=dataset.description,
                scope=dataset.scope,
                row_label=dataset.row_label,
                fields=[
                    ExportFieldPublic(
                        field=field.field,
                        label=field.label,
                        type=field.type,
                        sensitivity=field.sensitivity,
                        filter_operators=list(field.filter_operators),
                    )
                    for field in dataset.fields
                ],
            )
            for dataset in DATASETS.values()
        ],
        formats=[ExportFormat.CSV, ExportFormat.XLSX],
        limits={
            "max_columns": MAX_COLUMNS,
            "max_filters": 20,
            "max_rows": MAX_EXPORT_ROWS,
        },
    )


def _dataset(spec: CustomExportSpec) -> ExportDatasetDefinition:
    dataset = DATASETS.get(spec.dataset)
    if dataset is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Unknown export dataset: {spec.dataset}",
        )
    if dataset.scope == "gathering" and spec.popup_id is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Select a gathering for this export",
        )
    if dataset.scope == "organization" and spec.popup_id is not None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="This organization-level dataset does not accept a gathering",
        )
    return dataset


def _normalize_spec(
    spec: CustomExportSpec,
) -> tuple[CustomExportSpec, ExportDatasetDefinition, list[ExportFieldDefinition]]:
    dataset = _dataset(spec)
    fields_by_name = dataset.fields_by_name
    selected: list[ExportFieldDefinition] = []
    normalized_columns: list[ExportColumn] = []
    seen: set[str] = set()
    for column in spec.columns:
        field = fields_by_name.get(column.field)
        if field is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Field {column.field!r} is not available in {dataset.dataset}",
            )
        if column.field in seen:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Field {column.field!r} is selected more than once",
            )
        seen.add(column.field)
        selected.append(field)
        normalized_columns.append(
            ExportColumn(field=column.field, label=column.label or field.label)
        )

    normalized_filters: list[ExportFilter] = []
    for condition in spec.filters:
        field = fields_by_name.get(condition.field)
        if field is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Filter field {condition.field!r} is not available",
            )
        if condition.operator not in field.filter_operators:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    f"Operator {condition.operator!r} is not valid for "
                    f"{condition.field!r}"
                ),
            )
        if (
            condition.operator not in {"is_empty", "not_empty"}
            and condition.value is None
        ):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Filter {condition.field!r} requires a value",
            )
        if condition.operator == "in" and not isinstance(condition.value, list):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Filter {condition.field!r} with operator 'in' requires a list",
            )
        normalized_filters.append(condition)

    filename = spec.filename or dataset.dataset
    extension = spec.format.value
    if filename.lower().endswith(extension) and len(filename) > len(extension):
        filename = filename[: -len(extension)].rstrip(" ._-") or dataset.dataset
    normalized = spec.model_copy(
        update={
            "columns": normalized_columns,
            "filters": normalized_filters,
            "filename": filename,
        }
    )
    return normalized, dataset, selected


def _source_records(
    db: Session,
    dataset: ExportDatasetDefinition,
    popup_id: uuid.UUID | None,
) -> list[Any]:
    statement = select(dataset.model)
    if dataset.popup_mode == "column":
        statement = statement.where(dataset.model.popup_id == popup_id)
    elif dataset.popup_mode == "attendee_join":
        statement = statement.join(Attendees).where(Attendees.popup_id == popup_id)
    if dataset.exclude_deleted:
        statement = statement.where(dataset.model.deleted_at.is_(None))
    if dataset.option_factories:
        statement = statement.options(
            *(factory() for factory in dataset.option_factories)
        )
    statement = statement.order_by(dataset.model.id).limit(MAX_SOURCE_ROWS + 1)
    records = list(db.exec(statement).all())
    if len(records) > MAX_SOURCE_ROWS:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=(
                f"This export scans more than {MAX_SOURCE_ROWS:,} source records. "
                "Add narrower filters or choose a gathering-scoped dataset."
            ),
        )
    return records


def _is_empty(value: Any) -> bool:
    return value is None or value == "" or value == [] or value == ()


def _number(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _timestamp(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if not isinstance(value, str):
        return None
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=UTC)
    except ValueError:
        return None


def _equals(left: Any, right: Any, field_type: str) -> bool:
    if field_type == "number":
        return _number(left) == _number(right)
    if field_type == "boolean":
        expected = right if isinstance(right, bool) else str(right).lower() == "true"
        return left is expected
    return str(left).casefold() == str(right).casefold()


def _matches(value: Any, condition: ExportFilter, field: ExportFieldDefinition) -> bool:
    operator = condition.operator
    expected = condition.value
    if operator == "is_empty":
        return _is_empty(value)
    if operator == "not_empty":
        return not _is_empty(value)
    if operator == "eq":
        return not _is_empty(value) and _equals(value, expected, field.type)
    if operator == "neq":
        return _is_empty(value) or not _equals(value, expected, field.type)
    if operator == "in":
        if not isinstance(expected, list):
            return False
        return any(_equals(value, candidate, field.type) for candidate in expected)
    if operator in {"contains", "not_contains"}:
        contains = str(expected).casefold() in str(value or "").casefold()
        return contains if operator == "contains" else not contains
    if operator in {"before", "after"}:
        left = _timestamp(value)
        right = _timestamp(expected)
        if left is None or right is None:
            return False
        return left < right if operator == "before" else left > right
    if operator in {"gt", "gte", "lt", "lte"}:
        left = _number(value)
        right = _number(expected)
        if left is None or right is None:
            return False
        return {
            "gt": left > right,
            "gte": left >= right,
            "lt": left < right,
            "lte": left <= right,
        }[operator]
    return False


def _export_rows(
    db: Session,
    spec: CustomExportSpec,
) -> tuple[
    CustomExportSpec,
    ExportDatasetDefinition,
    list[ExportFieldDefinition],
    list[list[Any]],
]:
    normalized, dataset, selected = _normalize_spec(spec)
    fields_by_name = dataset.fields_by_name
    records = _source_records(db, dataset, normalized.popup_id)
    rows: list[list[Any]] = []
    for record in records:
        if any(
            not _matches(
                fields_by_name[condition.field].extractor(record),
                condition,
                fields_by_name[condition.field],
            )
            for condition in normalized.filters
        ):
            continue
        rows.append([field.extractor(record) for field in selected])
        if len(rows) > MAX_EXPORT_ROWS:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail=f"The export exceeds the {MAX_EXPORT_ROWS:,}-row limit",
            )
    return normalized, dataset, selected, rows


def _fingerprint(spec: CustomExportSpec) -> str:
    payload = json.dumps(
        spec.model_dump(mode="json"),
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(payload.encode()).hexdigest()[:16]


def preview_export(db: Session, spec: CustomExportSpec) -> ExportPreview:
    normalized, dataset, selected, rows = _export_rows(db, spec)
    sensitivities = {field.sensitivity for field in selected}
    warnings: list[str] = []
    if "pii" in sensitivities:
        warnings.append("This export contains personally identifiable information")
    if "financial" in sensitivities:
        warnings.append("This export contains financial information")
    if "security" in sensitivities:
        warnings.append("This export contains access or security information")
    if len(rows) >= 10_000:
        warnings.append(f"This is a large export with {len(rows):,} rows")

    extension = normalized.format.value
    filename = f"{normalized.filename}.{extension}"
    return ExportPreview(
        title=f"Export {dataset.label.lower()}",
        dataset=dataset.dataset,
        dataset_label=dataset.label,
        scope=dataset.scope,
        row_label=dataset.row_label,
        estimated_rows=len(rows),
        columns=[
            ExportPreviewColumn(
                field=field.field,
                label=column.label or field.label,
                type=field.type,
                sensitivity=field.sensitivity,
            )
            for column, field in zip(normalized.columns, selected, strict=True)
        ],
        filters=normalized.filters,
        warnings=warnings,
        format=normalized.format,
        filename=filename,
        spec=normalized,
        fingerprint=_fingerprint(normalized),
        generated_at=datetime.now(UTC),
    )


def _csv_bytes(headers: list[str], rows: list[list[Any]]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow([safe_tabular_cell(header) for header in headers])
    for row in rows:
        writer.writerow([safe_tabular_cell(value) for value in row])
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _xlsx_bytes(headers: list[str], rows: list[list[Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Export")
    sheet.freeze_panes = "A2"
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(len(header) + 2, 12), 40
        )
    header_cells = []
    for header in headers:
        cell = WriteOnlyCell(
            sheet, value=safe_tabular_cell(header, spreadsheet=True)
        )
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
        header_cells.append(cell)
    sheet.append(header_cells)
    for row in rows:
        sheet.append([safe_tabular_cell(value, spreadsheet=True) for value in row])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_export(
    db: Session,
    request_spec: CustomExportSpec,
    expected_fingerprint: str,
) -> tuple[bytes, str, str]:
    normalized, _dataset_definition, selected, rows = _export_rows(db, request_spec)
    fingerprint = _fingerprint(normalized)
    if fingerprint != expected_fingerprint:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="The export plan changed after it was previewed",
        )
    headers = [
        column.label or field.label
        for column, field in zip(normalized.columns, selected, strict=True)
    ]
    if normalized.format == ExportFormat.XLSX:
        content = _xlsx_bytes(headers, rows)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        content = _csv_bytes(headers, rows)
        media_type = "text/csv; charset=utf-8"
    return content, media_type, f"{normalized.filename}.{normalized.format.value}"
